import csv
import io
from pathlib import Path
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

class ReportsService:
    def __init__(self, repo):
        self.repo = repo

    def get_dashboard_summary(self,fecha_inicio=None, fecha_fin=None):
        return self.repo.dashboard_summary(fecha_inicio, fecha_fin)

    def generate_dashboard_csv(self,fecha_inicio=None, fecha_fin=None) -> str:
        data = self.get_dashboard_summary(fecha_inicio, fecha_fin)

        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(["Campo", "Valor"])
        writer.writerow(["Tareas totales", data["tareas_total"]])
        writer.writerow(["Tareas pendientes", data["tareas_pendientes"]])
        writer.writerow(["Tareas en progreso", data["tareas_en_progreso"]])
        writer.writerow(["Tareas completadas", data["tareas_completadas"]])
        writer.writerow(["Tareas canceladas", data["tareas_canceladas"]])
        writer.writerow(["Tareas vencidas", data["tareas_vencidas"]])
        writer.writerow(["Recursos totales", data["recursos_total"]])
        writer.writerow(["Recursos sin stock", data["recursos_sin_stock"]])
        writer.writerow(["Recursos en mantenimiento", data["recursos_en_mantenimiento"]])

        return output.getvalue()
    
    def generate_dashboard_excel(self,fecha_inicio=None, fecha_fin=None) -> bytes:
        data = self.get_dashboard_summary(fecha_inicio, fecha_fin)

        #wb = Workbook()
        #ws = wb.active
        #ws.title = "Dashboard"

        base_dir = Path(__file__).resolve().parent
        template_path = base_dir.parent / "templates" / "plantilla_dashboard.xlsx"
        #template_path = os.path.join("../templates", "plantilla_dashboard.xlsx")
        print('pathn del template ============>>>>>>>>>>>>>>>>>>>>>>>>>>>')
        print (template_path)

        wb = load_workbook(template_path)
        ws = wb["datos"]  # nombre de la hoja

        # Título
        ws["A1"] = "Reporte Dashboard"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Encabezados
        ws["A3"] = "Campo"
        ws["B3"] = "Valor"

        header_fill = PatternFill("solid", fgColor="D9EAD3")
        header_font = Font(bold=True)

        ws["A3"].fill = header_fill
        ws["B3"].fill = header_fill
        ws["A3"].font = header_font
        ws["B3"].font = header_font

        rows = [
            ("Tareas totales", data["tareas_total"]),
            ("Tareas pendientes", data["tareas_pendientes"]),
            ("Tareas en progreso", data["tareas_en_progreso"]),
            ("Tareas completadas", data["tareas_completadas"]),
            ("Tareas canceladas", data["tareas_canceladas"]),
            ("Tareas vencidas", data["tareas_vencidas"]),
            ("Recursos totales", data["recursos_total"]),
            ("Recursos sin stock", data["recursos_sin_stock"]),
            ("Recursos en mantenimiento", data["recursos_en_mantenimiento"]),
        ]

        start_row = 4
        for i, (campo, valor) in enumerate(rows, start=start_row):
            ws[f"A{i}"] = campo
            ws[f"B{i}"] = valor

        # Anchos de columna
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()